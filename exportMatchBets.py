import json
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from models.db import getConnection


class MatchBetsExporter:
    def __init__(self):
        self.db = getConnection()
        self.output_file = "MatchBetsExport.xlsx"
        print("Initialized MatchBetsExporter")

    def fetch_matches(self):
        with self.db.cursor() as cursor:
            cursor.execute("SELECT id, match_title FROM matches")
            return cursor.fetchall()

    def fetch_questions_for_match(self, match_id):
        with self.db.cursor() as cursor:
            cursor.execute(
                "SELECT id, question, options, correct_option FROM match_questions WHERE match_id = %s",
                (match_id,),
            )
            return {q["id"]: q for q in cursor.fetchall()}

    def fetch_bets_for_match(self, match_id):
        with self.db.cursor() as cursor:
            cursor.execute(
                """
                SELECT b.user_id, u.name, b.answers
                FROM match_bets b
                LEFT JOIN users u ON b.user_id = u.id
                WHERE b.match_id = %s
                """,
                (match_id,),
            )
            return cursor.fetchall()

    def get_option_details(self, options, option_id):
        """Return the full option dict matching the given option ID."""
        for opt in options:
            if str(opt["id"]) == str(option_id):
                return opt
        return None

    def calculate_points(self, amount, user_option_id, correct_option_id, options):
        correct_option_detail = self.get_option_details(options, correct_option_id)
        if not correct_option_detail:
            return 0

        # If correct option is "void", return 0
        if str(correct_option_detail["option"]).strip().lower() == "void":
            return 0

        if str(user_option_id) == str(correct_option_id):
            odds = correct_option_detail.get("odds", 1)
            return round(amount * odds - amount, 2)
        else:
            return -amount

    def export_to_excel(self):
        print("Exporting match bets to Excel...")
        wb = Workbook()
        wb.remove(wb.active)

        matches = self.fetch_matches()

        for match in matches:
            match_id = match["id"]
            title = re.sub(r"[:\\/?*\[\]]", "_", match["match_title"])[:25]
            sheet_name = f"{match_id}_{title}"
            print(f"Processing Match: {sheet_name}")

            questions = self.fetch_questions_for_match(match_id)
            bets = self.fetch_bets_for_match(match_id)

            if not bets or not questions:
                print(f"Skipping match {match_id} due to no data.")
                continue

            ws = wb.create_sheet(title=sheet_name)
            ws.append(
                [
                    "User ID",
                    "Name",
                    "Question ID",
                    "Question",
                    "User Option",
                    "Correct Option",
                    "Points",
                ]
            )

            for bet in bets:
                try:
                    answers = json.loads(bet["answers"])
                except json.JSONDecodeError:
                    continue

                user_id = bet["user_id"]
                user_name = bet["name"]

                for qid_str, ans_data in answers.items():
                    qid = int(qid_str)
                    if qid not in questions:
                        continue

                    amount = ans_data.get("amount", 0)
                    user_option_id = ans_data.get("option")

                    if not user_option_id or not amount:
                        continue

                    question_data = questions[qid]
                    try:
                        options = json.loads(question_data["options"])
                    except Exception:
                        continue

                    correct_option_id = question_data["correct_option"]

                    user_option_detail = self.get_option_details(
                        options, user_option_id
                    )
                    correct_option_detail = self.get_option_details(
                        options, correct_option_id
                    )

                    user_option_text = (
                        user_option_detail["option"] if user_option_detail else "N/A"
                    )
                    correct_option_text = (
                        correct_option_detail["option"]
                        if correct_option_detail
                        else "N/A"
                    )

                    points = self.calculate_points(
                        amount, user_option_id, correct_option_id, options
                    )

                    ws.append(
                        [
                            user_id,
                            user_name,
                            qid,
                            question_data["question"],
                            user_option_text,
                            correct_option_text,
                            points,
                        ]
                    )

            # Auto-adjust column widths
            for col in ws.columns:
                max_length = max(len(str(cell.value) or "") for cell in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = (
                    max_length + 2
                )

        wb.save(self.output_file)
        print(f"\n✅ Export completed: {self.output_file}")

    def close(self):
        print("Closing DB connection.")
        self.db.close()


def main():
    exporter = MatchBetsExporter()
    try:
        exporter.export_to_excel()
    finally:
        exporter.close()


if __name__ == "__main__":
    main()
