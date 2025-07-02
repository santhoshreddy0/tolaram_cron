import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from models.db import getConnection


class RoundUserProfitLossExporter:
    def __init__(self):
        self.db = getConnection()
        self.output_file = "UserRoundSummary.xlsx"
        print("Initialized RoundUserProfitLossExporter")

    def fetch_rounds(self):
        with self.db.cursor() as cursor:
            cursor.execute("SELECT id, round_name FROM rounds")
            return cursor.fetchall()

    def fetch_user_points_for_round(self, round_id):
        with self.db.cursor() as cursor:
            cursor.execute("""
                SELECT b.user_id, u.email, b.points
                FROM round_bets b
                LEFT JOIN users u ON b.user_id = u.id
                WHERE b.round_id = %s
            """, (round_id,))
            return cursor.fetchall()

    def get_result_from_points(self, points):
        if points > 0:
            return "Profit"
        elif points < 0:
            return "Loss"
        else:
            return "No Result"

    def export_to_excel(self):
        print("Exporting user profit/loss summary for rounds to Excel...")
        wb = Workbook()
        wb.remove(wb.active)

        rounds = self.fetch_rounds()

        for rnd in rounds:
            round_id = rnd["id"]
            title = re.sub(r"[:\\/?*\[\]]", "_", rnd["round_name"])[:25]
            sheet_name = f"{round_id}_{title}"
            print(f"Processing Round: {sheet_name}")

            user_bets = self.fetch_user_points_for_round(round_id)

            if not user_bets:
                print(f"Skipping round {round_id} due to no bets.")
                continue

            ws = wb.create_sheet(title=sheet_name)
            ws.append(["User ID", "Email", "Points", "Result"])

            for row in user_bets:
                user_id = row["user_id"]
                user_email = row["email"]
                points = row["points"]
                result = self.get_result_from_points(points)

                ws.append([user_id, user_email, points, result])

            # Auto-adjust column widths
            for col in ws.columns:
                max_length = max(len(str(cell.value) or "") for cell in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

        wb.save(self.output_file)
        print(f"\n Export completed: {self.output_file}")

    def close(self):
        print("Closing DB connection.")
        self.db.close()


def main():
    exporter = RoundUserProfitLossExporter()
    try:
        exporter.export_to_excel()
    finally:
        exporter.close()


if __name__ == "__main__":
    main()
