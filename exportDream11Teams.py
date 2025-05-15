import re
import datetime
from models.db import getConnection
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class Dream11TeamExporter:
    def __init__(self):
        self.db = getConnection()
        self.output_file = "Super12Teams.xlsx"
        print("Initialized Dream11TeamExporter")

    def fetch_all_user_ids(self):
        print("Fetching all user IDs...")
        with self.db.cursor() as cursor:
            cursor.execute("SELECT DISTINCT user_id FROM dream11_players")
            user_ids = [row["user_id"] for row in cursor.fetchall()]
        print(f"Found {len(user_ids)} unique user IDs")
        return user_ids

    def fetch_user_team_data(self, user_id):
        print(f"Fetching team data for user ID: {user_id}")
        sql = """
            SELECT
                u.email,
                p.name AS player_name,
                t.team_name AS player_team,
                dp.role_type AS player_role,
                IFNULL(SUM(mpm.points), 0) * 
                  CASE
                    WHEN dp.role_type = 'captain' THEN 2
                    WHEN dp.role_type = 'vice-captain' THEN 1.5
                    ELSE 1
                  END AS points
            FROM dream11_players dp
            LEFT JOIN match_player_mapping mpm ON dp.player_id = mpm.player_id
            LEFT JOIN players p ON p.id = dp.player_id
            LEFT JOIN teams t ON p.team_id = t.id
            LEFT JOIN users u ON dp.user_id = u.id
            WHERE dp.user_id = %s
            GROUP BY dp.player_id, dp.role_type
        """
        with self.db.cursor() as cursor:
            cursor.execute(sql, (user_id,))
            team_data = cursor.fetchall()
        print(f"Fetched {len(team_data)} players for user ID {user_id}")
        return team_data

    def sanitize_sheet_name(self, email, fallback_name):
        if not email:
            return fallback_name

        # Remove invalid characters and limit to 31 chars
        name = re.sub(r"[:\\/?*\[\]]", "_", email)
        return name[:31]

    def export_to_excel(self):
        print("Starting export to Excel...")
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        user_ids = self.fetch_all_user_ids()
        sheet_names = set()

        for user_id in user_ids:
            print(f"\nProcessing user ID: {user_id}")
            team_data = self.fetch_user_team_data(user_id)

            if not team_data:
                print(f"No team data found for user ID {user_id}. Skipping...")
                continue

            raw_email = team_data[0]["email"]
            base_name = self.sanitize_sheet_name(raw_email, f"user_{user_id}")

            # Ensure unique sheet name
            sheet_name = base_name
            counter = 1
            while sheet_name in sheet_names:
                suffix = f"_{counter}"
                max_length = 31 - len(suffix)
                sheet_name = base_name[:max_length] + suffix
                counter += 1
            sheet_names.add(sheet_name)

            print(f"Creating sheet: {sheet_name}")
            ws = wb.create_sheet(title=sheet_name)
            ws.append(["Player Name", "Team Name", "Role", "Points"])

            for row in team_data:
                ws.append(
                    [
                        row["player_name"],
                        row["player_team"],
                        row["player_role"],
                        float(row["points"]),
                    ]
                )

            # Auto-size columns
            for col in ws.columns:
                max_length = max(
                    len(str(cell.value)) if cell.value else 0 for cell in col
                )
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max_length + 2

        wb.save(self.output_file)
        print(f"\n✅ Export completed: {self.output_file}")

    def close(self):
        print("Closing database connection...")
        self.db.close()


def main():
    print("Starting the export process...")
    exporter = Dream11TeamExporter()
    try:
        exporter.export_to_excel()
    finally:
        exporter.close()


if __name__ == "__main__":
    main()
