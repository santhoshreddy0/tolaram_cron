import json
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from models.db import getConnection


class RoundBetsExporter:
    def __init__(self):
        self.db = getConnection()
        self.output_file = "RoundBetsExport.xlsx"
        print("Initialized RoundBetsExporter")

    def fetch_rounds(self):
        with self.db.cursor() as cursor:
            cursor.execute("SELECT id, round_name FROM rounds")
            return cursor.fetchall()

    def fetch_questions_for_round(self, round_id):
        with self.db.cursor() as cursor:
            cursor.execute(
                """
                SELECT id, question, options, correct_option 
                FROM round_questions 
                WHERE round_id = %s
                """,
                (round_id,),
            )
            return {q["id"]: q for q in cursor.fetchall()}

    def fetch_bets_for_round(self, round_id):
        with self.db.cursor() as cursor:
            cursor.execute(
                """
                SELECT b.user_id, u.name, b.answers
                FROM round_bets b
                LEFT JOIN users u ON b.user_id = u.id
                WHERE b.round_id = %s
                """,
                (round_id,),
            )
            return cursor.fetchall()

    def calculate_points(self, amount, user_option_id, correct_option_id, options_map):
        if correct_option_id is None:
            return 0

        user_opt_data = options_map.get(str(user_option_id))
        correct_opt_data = options_map.get(str(correct_option_id))

        if not user_opt_data or not correct_opt_data:
            return 0

        # If correct option is marked as "void"
        if correct_opt_data["option"].lower() == "void":
            return 0

        if str(user_option_id) == str(correct_option_id):
            odds = user_opt_data.get("odds", 1)
            return round(amount * odds - amount, 2)
        else:
            return -amount

    def export_to_excel(self):
        print("Exporting round bets to Excel...")
        wb = Workbook()
        wb.remove(wb.active)

        rounds = self.fetch_rounds()

        for round_ in rounds:
            round_id = round_["id"]
            title = re.sub(r"[:\\/?*\[\]]", "_", round_["round_name"])[:25]
            sheet_name = f"{round_id}_{title}"
            print(f"Processing Round: {sheet_name}")

            questions = self.fetch_questions_for_round(round_id)
            bets = self.fetch_bets_for_round(round_id)

            if not bets or not questions:
                print(f"Skipping round {round_id} due to no data.")
                continue

            ws = wb.create_sheet(title=sheet_name)
            ws.append(
                [
                    "User ID",
                    "Name",
                    "Question ID",
                    "Question",
                    "User Option",
                    "Correct Option",
                    "Points",
                ]
            )

            for bet in bets:
                try:
                    answers = json.loads(bet["answers"])
                except json.JSONDecodeError:
                    continue

                user_id = bet["user_id"]
                user_name = bet["name"]

                for qid_str, ans_data in answers.items():
                    if not isinstance(ans_data, dict):
                        continue

                    qid = int(qid_str)
                    if qid not in questions:
                        continue

                    question_data = questions[qid]
                    try:
                        options = json.loads(question_data["options"])
                    except json.JSONDecodeError:
                        continue

                    options_map = {str(opt["id"]): opt for opt in options}

                    user_option_id = ans_data.get("option")
                    amount = ans_data.get("amount", 0)
                    if amount == 0 or user_option_id is None:
                        continue

                    user_option_label = options_map.get(str(user_option_id), {}).get(
                        "option", ""
                    )
                    correct_option_id = question_data["correct_option"]
                    correct_option_label = options_map.get(
                        str(correct_option_id), {}
                    ).get("option", "")

                    points = self.calculate_points(
                        amount, user_option_id, correct_option_id, options_map
                    )

                    ws.append(
                        [
                            user_id,
                            user_name,
                            qid,
                            question_data["question"],
                            user_option_label,
                            correct_option_label,
                            points,
                        ]
                    )

            # Auto-adjust column widths
            for col in ws.columns:
                max_length = max(len(str(cell.value) or "") for cell in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = (
                    max_length + 2
                )

        wb.save(self.output_file)
        print(f"\n✅ Export completed: {self.output_file}")

    def close(self):
        print("Closing DB connection.")
        self.db.close()


def main():
    exporter = RoundBetsExporter()
    try:
        exporter.export_to_excel()
    finally:
        exporter.close()


if __name__ == "__main__":
    main()
