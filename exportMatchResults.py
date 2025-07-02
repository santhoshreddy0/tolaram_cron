import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from models.db import getConnection


class MatchUserProfitLossExporter:
    def __init__(self):
        self.db = getConnection()
        self.output_file = "UserMatchSummary.xlsx"
        print("Initialized MatchUserProfitLossExporter")

    def fetch_matches(self):
        with self.db.cursor() as cursor:
            cursor.execute("SELECT id, match_title FROM matches")
            return cursor.fetchall()

    def fetch_user_points_for_match(self, match_id):
        with self.db.cursor() as cursor:
            cursor.execute("""
                SELECT b.user_id, u.email, b.points
                FROM match_bets b
                LEFT JOIN users u ON b.user_id = u.id
                WHERE b.match_id = %s
            """, (match_id,))
            return cursor.fetchall()

    def get_result_from_points(self, points):
        if points > 0:
            return "Profit"
        elif points < 0:
            return "Loss"
        else:
            return "No Result"

    def export_to_excel(self):
        print("Exporting user profit/loss summary to Excel...")
        wb = Workbook()
        wb.remove(wb.active)

        matches = self.fetch_matches()

        for match in matches:
            match_id = match["id"]
            title = re.sub(r"[:\\/?*\[\]]", "_", match["match_title"])[:25]
            sheet_name = f"{match_id}_{title}"
            print(f"Processing Match: {sheet_name}")

            user_bets = self.fetch_user_points_for_match(match_id)

            if not user_bets:
                print(f"Skipping match {match_id} due to no bets.")
                continue

            ws = wb.create_sheet(title=sheet_name)
            ws.append(["User ID", "Email", "Points", "Result"])

            for row in user_bets:
                user_id = row["user_id"]
                user_email = row["email"]
                points = row["points"]
                result = self.get_result_from_points(points)

                ws.append([user_id, user_email, points, result])

            # Auto-adjust column widths
            for col in ws.columns:
                max_length = max(len(str(cell.value) or "") for cell in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

        wb.save(self.output_file)
        print(f"\n✅ Export completed: {self.output_file}")

    def close(self):
        print("Closing DB connection.")
        self.db.close()


def main():
    exporter = MatchUserProfitLossExporter()
    try:
        exporter.export_to_excel()
    finally:
        exporter.close()


if __name__ == "__main__":
    main()
